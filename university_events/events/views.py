import json
import re
from django.utils import timezone
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from .models import Event, Registration, EventType, MainImage, SubCategory, EventImage
from .forms import RegistrationForm
import openpyxl
from openpyxl.styles import Font
import markdown
from django.db.models import Q
from django.core.paginator import Paginator

def index(request):
    event_type_id = request.GET.get('event_type', None)
    subcategory_id = request.GET.get('subcategory', None)

    events = Event.objects.filter(
        Q(date__gte=timezone.now().date())
    ).order_by("date", "start_time")

  
    main_image = MainImage.objects.first()
    event_types = EventType.objects.all()

    if event_type_id:
        events = events.filter(event_type_id=event_type_id)

    if subcategory_id:
        events = events.filter(subcategory_id=subcategory_id)

    entertainment_event_type = EventType.objects.filter(name="Развлекательные мероприятия").first()
    subcategories = SubCategory.objects.filter(event_type=entertainment_event_type) if entertainment_event_type else []

    return render(request, "index.html", {
        "events": events,
        "event_types": event_types,
        "selected_event_type": event_type_id,
        "subcategories": subcategories,
        "selected_subcategory": subcategory_id,
        'main_image': main_image,
    })


def calendar_view(request):
    """ Страница календаря с событиями """
    events = Event.objects.all().values("id", "title", "date", "event_type__color")
    events_list = []
    for event in events:
        events_list.append({
            "id": event["id"],
            "title": event["title"],
            "start": str(event["date"]),  
            "color": event["event_type__color"] if event["event_type__color"] else "#007bff"
        })

    events_json = json.dumps(events_list)

    return render(request, "calendar.html", {"events_json": events_json})


def event_detail(request, event_id):
    """ Детальная страница мероприятия """
    event = get_object_or_404(Event, id=event_id)

  
    event.description = markdown.markdown(event.description)

    error_message = None  

    if request.method == "POST":
        form = RegistrationForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']

            
            existing_registration = Registration.objects.filter(event=event, email=email).first()
            if existing_registration:
                error_message = "Вы уже зарегистрированы на это событие с этим email."
            else:
                
                registration = form.save(commit=False)
                registration.event = event
                registration.save()

                return redirect("success")

        else:
            
            error_message = "Ошибка при регистрации. Пожалуйста, проверьте данные и попробуйте снова."

    else:
        form = RegistrationForm()

    return render(request, "event_detail.html", {
        "event": event,
        "form": form,
        "error_message": error_message,  
    })


import re

def my_events(request):
    phone = request.GET.get('phone', '')
    email = request.GET.get('email', '')
    
    
    clean_phone = re.sub(r'\D', '', phone) if phone else ''
    
    if clean_phone:
        
        phone_for_search = clean_phone[-10:] if len(clean_phone) > 10 else clean_phone
        registrations = Registration.objects.filter(
            phone_number__contains=phone_for_search
        )
    elif email:
        registrations = Registration.objects.filter(email=email)
    else:
        registrations = Registration.objects.none()
    
    return render(request, "my_events.html", {
        'registrations': registrations,
        'search_phone': phone,
        'search_email': email
    })



def success(request):
    """ Страница успешной регистрации """
    return render(request, "success.html")

@staff_member_required
def admin_settings(request):
    """Страница настроек для администратора"""
    if not request.user.is_staff:
        return redirect('index')  

    return render(request, 'admin_settings.html')

@staff_member_required
def export_registrations_to_excel(request):
    """Экспорт данных о регистрациях в Excel с сортировкой по мероприятиям"""
    registrations = Registration.objects.select_related("event").order_by("event__title", "last_name")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Регистрации"

    
    headers = ["Имя", "Фамилия", "Email", "Телефон", "Событие", "Дата регистрации"]

    header_font = Font(bold=True)

    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font

    
    for row_num, reg in enumerate(registrations, start=2):
        sheet.cell(row=row_num, column=1, value=reg.first_name)
        sheet.cell(row=row_num, column=2, value=reg.last_name)
        sheet.cell(row=row_num, column=3, value=reg.email)
        sheet.cell(row=row_num, column=4, value=reg.phone_number)
        sheet.cell(row=row_num, column=5, value=reg.event.title)

        created_at_display = reg.created_at.strftime("%d-%m-%Y %H:%M") if reg.created_at else "Не указано"
        sheet.cell(row=row_num, column=6, value=created_at_display)

   
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells if cell.value)
        sheet.column_dimensions[column_cells[0].column_letter].width = length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=registrations.xlsx"
    workbook.save(response)

    return response


def gallery_view(request):
    past_events = Event.objects.filter(date__lt=timezone.now().date()).order_by("-date")
    return render(request, "gallery.html", {"past_events": past_events})

def past_event_detail(request, event_id):
    event = get_object_or_404(Event, id=event_id)


    report_html = markdown.markdown(event.report) if event.report else ""

    images = event.images.all()

    return render(request, "past_event_detail.html", {
        "event": event,
        "report_html": report_html,
        "images": images,
    })




def gallery_view(request):
    past_events = Event.objects.filter(date__lt=timezone.now()).order_by('-date')
    paginator = Paginator(past_events, 6)  
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'gallery.html', {
        'past_events': page_obj  
    })
